#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
GPW SCREENER - Signal Aggregation & Investment Thesis Generator
================================================================================

Agreguje wyniki ze wszystkich modeli i generuje:
- Consensus ranking oparty na sygnałach
- Investment thesis dla każdej spółki
- Heatmapę flag
- Best-of w kategoriach

UŻYCIE:
    python app.py                    # Generuj wyniki_ostateczne.xlsx
    python app.py --status           # Status systemu
    python app.py --top N            # Pokaż TOP N w konsoli
    python app.py --ticker XYZ       # Szczegóły dla spółki

================================================================================
"""

import os
import sys
import argparse
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Ścieżki
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MAIN_DIR = os.path.join(BASE_DIR, "main")
LATEST_FILE = os.path.join(MAIN_DIR, "wyniki_latest.xlsx")
OUTPUT_FILE = os.path.join(MAIN_DIR, "wyniki_ostateczne.xlsx")

# Konfiguracja logowania
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# KONFIGURACJA SYGNAŁÓW I WAG
# =============================================================================

# Punkty za ranking w modelu
RANK_POINTS = {
    'top5': 5,
    'top10': 3,
    'top20': 1,
    'other': 0
}

# Kategoryzacja flag (które flagi oznaczają co)
FLAG_CATEGORIES = {
    'quality': ['Q'],           # Jakość biznesu
    'growth': ['G', 'R'],       # Wzrost
    'value': ['V', 'D'],        # Wycena
    'momentum': ['M', 'A'],     # Momentum
    'safety': ['S', 'B', 'L'],  # Bezpieczeństwo
    'cash': ['C'],              # Cash flow
    'turnaround': ['T'],        # Sygnały odbicia
    'warning': ['!', '?']       # Ostrzeżenia
}

# Odwrotne mapowanie: flaga -> kategoria
FLAG_TO_CATEGORY = {}
for cat, flags in FLAG_CATEGORIES.items():
    for f in flags:
        FLAG_TO_CATEGORY[f] = cat

# Wagi kategorii flag (do obliczania thesis clarity)
CATEGORY_WEIGHTS = {
    'quality': 1.5,
    'growth': 1.3,
    'value': 1.2,
    'momentum': 1.1,
    'safety': 1.4,
    'cash': 1.3,
    'turnaround': 1.0,
    'warning': -2.0  # Kara za ostrzeżenia
}

# Mapowanie modeli na główne "tematy"
MODEL_THEMES = {
    'Quality Growth': ['quality', 'growth'],
    'Turnaround': ['value', 'turnaround'],
    'Revenue Momentum & Safety': ['momentum', 'safety', 'growth'],
    'Cash Quality & Balance Sheet': ['cash', 'safety', 'quality'],
    'Cash Quality & Balance She': ['cash', 'safety', 'quality'],  # Skrócona nazwa
    'Quality Momentum': ['quality', 'momentum'],
    'Valuation Compression': ['value', 'momentum']
}


# =============================================================================
# KLASY DANYCH
# =============================================================================

class CompanySignals:
    """Agreguje sygnały dla pojedynczej spółki"""
    
    def __init__(self, ticker: str):
        self.ticker = ticker
        self.rynek = None
        self.appearances: Dict[str, dict] = {}  # model_name -> {rank, score, flags}
        
    def add_appearance(self, model_name: str, rank: int, score: float, flags: str, rynek: str = None):
        """Dodaje wystąpienie w modelu"""
        self.appearances[model_name] = {
            'rank': rank,
            'score': score,
            'flags': flags,
            'flags_list': self._parse_flags(flags)
        }
        if rynek and not self.rynek:
            self.rynek = rynek
    
    def _parse_flags(self, flags_str: str) -> List[str]:
        """Parsuje flagi ze stringa [Q][G][V] -> ['Q', 'G', 'V']"""
        if not flags_str or pd.isna(flags_str):
            return []
        import re
        return re.findall(r'\[([A-Z!?]+)\]', str(flags_str))
    
    @property
    def coverage(self) -> int:
        """W ilu modelach występuje"""
        return len(self.appearances)
    
    @property
    def elite_score(self) -> int:
        """Suma punktów za rankingi (TOP5=5, TOP10=3, TOP20=1)"""
        total = 0
        for app in self.appearances.values():
            rank = app['rank']
            if rank <= 5:
                total += RANK_POINTS['top5']
            elif rank <= 10:
                total += RANK_POINTS['top10']
            elif rank <= 20:
                total += RANK_POINTS['top20']
        return total
    
    @property
    def top5_count(self) -> int:
        """Ile razy w TOP5"""
        return sum(1 for app in self.appearances.values() if app['rank'] <= 5)
    
    @property
    def top10_count(self) -> int:
        """Ile razy w TOP10"""
        return sum(1 for app in self.appearances.values() if app['rank'] <= 10)
    
    @property
    def all_flags(self) -> List[str]:
        """Wszystkie flagi ze wszystkich modeli"""
        flags = []
        for app in self.appearances.values():
            flags.extend(app['flags_list'])
        return flags
    
    @property
    def unique_flags(self) -> Set[str]:
        """Unikalne flagi"""
        return set(self.all_flags)
    
    @property
    def warning_count(self) -> int:
        """Ile flag ostrzegawczych"""
        return sum(1 for f in self.all_flags if f in ['!', '?'])
    
    @property
    def positive_flag_count(self) -> int:
        """Ile pozytywnych flag (bez ostrzeżeń)"""
        return sum(1 for f in self.all_flags if f not in ['!', '?'])
    
    def get_flag_density(self) -> float:
        """Średnia liczba pozytywnych flag na model"""
        if self.coverage == 0:
            return 0
        return self.positive_flag_count / self.coverage
    
    def get_category_strength(self) -> Dict[str, float]:
        """Siła w każdej kategorii (ile flag z tej kategorii)"""
        strength = defaultdict(float)
        for flag in self.all_flags:
            if flag in FLAG_TO_CATEGORY:
                cat = FLAG_TO_CATEGORY[flag]
                strength[cat] += CATEGORY_WEIGHTS.get(cat, 1.0)
        return dict(strength)
    
    def get_dominant_category(self) -> Tuple[str, float]:
        """Dominująca kategoria (najsilniejsza)"""
        strength = self.get_category_strength()
        if not strength:
            return ('unknown', 0)
        # Usuń warning z rozważań
        strength.pop('warning', None)
        if not strength:
            return ('unknown', 0)
        dominant = max(strength.items(), key=lambda x: x[1])
        return dominant
    
    def calculate_signal_strength(self) -> float:
        """
        Główna metryka - Signal Strength
        Kombinacja: elite_score + flag_density + coverage_bonus - warnings
        """
        # Bazowy score z rankingów
        base = self.elite_score
        
        # Bonus za gęstość flag
        flag_bonus = self.get_flag_density() * 3
        
        # Bonus za coverage (ale nie liniowy - diminishing returns)
        coverage_bonus = min(self.coverage, 4) * 1.5
        
        # Kara za ostrzeżenia
        warning_penalty = self.warning_count * 2
        
        # Bonus za spójność (wiele flag z tej samej kategorii)
        dominant_cat, dominant_strength = self.get_dominant_category()
        consistency_bonus = min(dominant_strength, 5) * 0.5
        
        return base + flag_bonus + coverage_bonus - warning_penalty + consistency_bonus
    
    def generate_thesis(self) -> str:
        """Generuje investment thesis dla spółki"""
        if self.coverage == 0:
            return "Brak danych"
        
        parts = []
        
        # Główna kategoria
        dominant_cat, strength = self.get_dominant_category()
        category_names = {
            'quality': 'QUALITY',
            'growth': 'GROWTH',
            'value': 'VALUE',
            'momentum': 'MOMENTUM',
            'safety': 'DEFENSIVE',
            'cash': 'CASH KING',
            'turnaround': 'TURNAROUND',
            'unknown': 'MIXED'
        }
        parts.append(category_names.get(dominant_cat, 'MIXED'))
        
        # Siła przekonania
        signal_str = self.calculate_signal_strength()
        if signal_str >= 15:
            conviction = "Strong"
        elif signal_str >= 10:
            conviction = "Medium"
        else:
            conviction = "Weak"
        
        # Szczegóły
        details = []
        
        # TOP appearances
        if self.top5_count >= 2:
            details.append(f"TOP5 w {self.top5_count} modelach")
        elif self.top5_count == 1:
            top5_models = [m for m, a in self.appearances.items() if a['rank'] <= 5]
            details.append(f"TOP5: {top5_models[0][:15]}")
        
        # Kluczowe flagi
        cat_strength = self.get_category_strength()
        cat_strength.pop('warning', None)
        
        # Top 2 kategorie
        sorted_cats = sorted(cat_strength.items(), key=lambda x: -x[1])[:2]
        flag_desc = {
            'quality': 'wysoka jakość',
            'growth': 'wzrost',
            'value': 'niska wycena',
            'momentum': 'momentum',
            'safety': 'bezpieczeństwo',
            'cash': 'cash flow',
            'turnaround': 'sygnały odbicia'
        }
        for cat, _ in sorted_cats:
            if cat in flag_desc:
                details.append(flag_desc[cat])
        
        # Ostrzeżenia
        if self.warning_count > 0:
            details.append(f"⚠️ {self.warning_count}x warning")
        
        # Złóż thesis
        thesis = f"{parts[0]} ({conviction})"
        if details:
            thesis += ": " + ", ".join(details)
        
        return thesis
    
    def get_models_summary(self) -> str:
        """Podsumowanie w jakich modelach występuje"""
        if not self.appearances:
            return ""
        summaries = []
        for model, data in sorted(self.appearances.items(), key=lambda x: x[1]['rank']):
            short_model = model[:12]
            summaries.append(f"{short_model}:#{data['rank']}")
        return " | ".join(summaries)


# =============================================================================
# FUNKCJE GŁÓWNE
# =============================================================================

def load_results() -> Dict[str, pd.DataFrame]:
    """Wczytuje wyniki z wyniki_latest.xlsx"""
    if not os.path.exists(LATEST_FILE):
        logger.error(f"❌ Brak pliku: {LATEST_FILE}")
        logger.error("   Uruchom najpierw: python run.py")
        return {}
    
    logger.info(f"📂 Wczytuję: {LATEST_FILE}")
    xl = pd.ExcelFile(LATEST_FILE)
    
    results = {}
    for sheet in xl.sheet_names:
        if sheet == "PODSUMOWANIE":
            continue
        df = pd.read_excel(xl, sheet_name=sheet)
        if 'Ticker' in df.columns and 'Rank' in df.columns:
            results[sheet] = df
            logger.info(f"   ✅ {sheet}: {len(df)} spółek")
    
    return results


def aggregate_signals(results: Dict[str, pd.DataFrame]) -> Dict[str, CompanySignals]:
    """Agreguje sygnały ze wszystkich modeli"""
    companies: Dict[str, CompanySignals] = {}
    
    for model_name, df in results.items():
        for _, row in df.iterrows():
            ticker = row.get('Ticker')
            if not ticker or pd.isna(ticker):
                continue
            
            if ticker not in companies:
                companies[ticker] = CompanySignals(ticker)
            
            companies[ticker].add_appearance(
                model_name=model_name,
                rank=int(row.get('Rank', 999)),
                score=float(row.get('Total', 0)),
                flags=str(row.get('Flags', '')),
                rynek=row.get('Rynek')
            )
    
    return companies


def create_consensus_df(companies: Dict[str, CompanySignals]) -> pd.DataFrame:
    """Tworzy DataFrame z consensus rankingiem"""
    rows = []
    
    for ticker, signals in companies.items():
        rows.append({
            'Ticker': ticker,
            'Rynek': signals.rynek or '',
            'Signal_Strength': round(signals.calculate_signal_strength(), 1),
            'Coverage': f"{signals.coverage}/{len(MODEL_THEMES)}",
            'Elite_Score': signals.elite_score,
            'TOP5_Count': signals.top5_count,
            'TOP10_Count': signals.top10_count,
            'Positive_Flags': signals.positive_flag_count,
            'Warnings': signals.warning_count,
            'Unique_Flags': ''.join(f'[{f}]' for f in sorted(signals.unique_flags)),
            'Investment_Thesis': signals.generate_thesis(),
            'Models': signals.get_models_summary()
        })
    
    df = pd.DataFrame(rows)
    df = df.sort_values('Signal_Strength', ascending=False).reset_index(drop=True)
    df.insert(0, 'Rank', range(1, len(df) + 1))
    
    return df


def create_flag_heatmap(companies: Dict[str, CompanySignals], top_n: int = 30) -> pd.DataFrame:
    """Tworzy heatmapę flag dla TOP N spółek"""
    # Sortuj po signal strength
    sorted_companies = sorted(
        companies.values(),
        key=lambda x: x.calculate_signal_strength(),
        reverse=True
    )[:top_n]
    
    # Wszystkie możliwe flagi
    all_flags = ['Q', 'G', 'V', 'M', 'A', 'R', 'S', 'B', 'L', 'C', 'D', 'T', '!', '?']
    
    rows = []
    for signals in sorted_companies:
        row = {'Ticker': signals.ticker}
        flag_counts = defaultdict(int)
        for f in signals.all_flags:
            flag_counts[f] += 1
        for flag in all_flags:
            row[f'[{flag}]'] = flag_counts.get(flag, 0)
        row['Total_Flags'] = signals.positive_flag_count
        rows.append(row)
    
    return pd.DataFrame(rows)


def create_best_of(companies: Dict[str, CompanySignals]) -> pd.DataFrame:
    """Tworzy ranking best-of w kategoriach"""
    categories = {
        'QUALITY': ['quality'],
        'GROWTH': ['growth'],
        'VALUE': ['value'],
        'MOMENTUM': ['momentum'],
        'SAFETY': ['safety'],
        'CASH': ['cash'],
        'TURNAROUND': ['turnaround']
    }
    
    rows = []
    for cat_name, cat_flags in categories.items():
        # Znajdź spółki z najsilniejszymi flagami w tej kategorii
        cat_scores = []
        for ticker, signals in companies.items():
            strength = signals.get_category_strength()
            cat_score = sum(strength.get(cf, 0) for cf in cat_flags)
            if cat_score > 0:
                cat_scores.append((ticker, cat_score, signals))
        
        # TOP 3 w kategorii
        cat_scores.sort(key=lambda x: -x[1])
        for i, (ticker, score, signals) in enumerate(cat_scores[:3], 1):
            rows.append({
                'Category': cat_name,
                'Rank': i,
                'Ticker': ticker,
                'Category_Score': round(score, 1),
                'Signal_Strength': round(signals.calculate_signal_strength(), 1),
                'Flags': ''.join(f'[{f}]' for f in sorted(signals.unique_flags))
            })
    
    return pd.DataFrame(rows)


def create_profiles(companies: Dict[str, CompanySignals], results: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Tworzy szczegółowe profile spółek"""
    rows = []
    
    # Sortuj po signal strength
    sorted_tickers = sorted(
        companies.keys(),
        key=lambda t: companies[t].calculate_signal_strength(),
        reverse=True
    )
    
    for ticker in sorted_tickers:
        signals = companies[ticker]
        
        # Szczegóły z każdego modelu
        for model_name, app_data in signals.appearances.items():
            rows.append({
                'Ticker': ticker,
                'Model': model_name,
                'Rank_in_Model': app_data['rank'],
                'Score_in_Model': round(app_data['score'], 1),
                'Flags_in_Model': app_data['flags'],
                'Signal_Strength': round(signals.calculate_signal_strength(), 1)
            })
    
    return pd.DataFrame(rows)


def save_excel(consensus_df: pd.DataFrame, heatmap_df: pd.DataFrame,
               best_of_df: pd.DataFrame, profiles_df: pd.DataFrame,
               output_path: str):
    """Zapisuje wyniki do Excel"""
    logger.info(f"\n📊 Zapisuję: {output_path}")
    
    wb = Workbook()
    
    # Style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    top3_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    top10_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def style_sheet(ws, df, highlight_top=True):
        """Stylizuje arkusz"""
        # Nagłówki
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Dane
        for row_idx, row_data in df.iterrows():
            excel_row = row_idx + 2
            for col_idx, col_name in enumerate(df.columns, 1):
                value = row_data[col_name]
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = thin_border
                
                # Kolorowanie
                if highlight_top and 'Rank' in df.columns:
                    rank = row_data.get('Rank', 999)
                    if rank <= 3:
                        cell.fill = top3_fill
                    elif rank <= 10:
                        cell.fill = top10_fill
                
                # Warning highlighting
                if col_name == 'Warnings' and value > 0:
                    cell.fill = warning_fill
        
        # Auto-width
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
    
    # === ARKUSZ 1: CONSENSUS ===
    ws_consensus = wb.active
    ws_consensus.title = "CONSENSUS"
    
    # Tytuł
    ws_consensus['A1'] = "GPW SCREENER - Consensus Ranking"
    ws_consensus['A1'].font = Font(bold=True, size=16)
    ws_consensus.merge_cells('A1:F1')
    ws_consensus['A2'] = f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_consensus['A2'].font = Font(italic=True, color="666666")
    
    # Dane od wiersza 4
    for col_idx, col_name in enumerate(consensus_df.columns, 1):
        cell = ws_consensus.cell(row=4, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, row_data in consensus_df.iterrows():
        excel_row = row_idx + 5
        for col_idx, col_name in enumerate(consensus_df.columns, 1):
            value = row_data[col_name]
            cell = ws_consensus.cell(row=excel_row, column=col_idx, value=value)
            cell.border = thin_border
            
            rank = row_data.get('Rank', 999)
            if rank <= 3:
                cell.fill = top3_fill
            elif rank <= 10:
                cell.fill = top10_fill
            
            if col_name == 'Warnings' and value > 0:
                cell.fill = warning_fill
    
    # Szerokości kolumn
    col_widths = {'Rank': 6, 'Ticker': 10, 'Rynek': 8, 'Signal_Strength': 14, 
                  'Coverage': 10, 'Elite_Score': 11, 'TOP5_Count': 11, 'TOP10_Count': 11,
                  'Positive_Flags': 13, 'Warnings': 10, 'Unique_Flags': 20,
                  'Investment_Thesis': 50, 'Models': 60}
    for col_idx, col_name in enumerate(consensus_df.columns, 1):
        ws_consensus.column_dimensions[ws_consensus.cell(row=4, column=col_idx).column_letter].width = col_widths.get(col_name, 12)
    
    # === ARKUSZ 2: FLAG HEATMAP ===
    ws_heatmap = wb.create_sheet("FLAG_HEATMAP")
    style_sheet(ws_heatmap, heatmap_df, highlight_top=False)
    
    # Koloruj komórki z flagami
    for row_idx in range(len(heatmap_df)):
        excel_row = row_idx + 2
        for col_idx in range(2, len(heatmap_df.columns)):  # Pomiń Ticker
            cell = ws_heatmap.cell(row=excel_row, column=col_idx)
            if cell.value and cell.value > 0:
                if cell.value >= 3:
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif cell.value >= 2:
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # === ARKUSZ 3: BEST OF ===
    ws_bestof = wb.create_sheet("BEST_OF")
    style_sheet(ws_bestof, best_of_df, highlight_top=False)
    
    # Koloruj kategorie
    category_colors = {
        'QUALITY': "4472C4",
        'GROWTH': "70AD47",
        'VALUE': "ED7D31",
        'MOMENTUM': "7030A0",
        'SAFETY': "00B0F0",
        'CASH': "FFC000",
        'TURNAROUND': "FF0000"
    }
    for row_idx, row_data in best_of_df.iterrows():
        excel_row = row_idx + 2
        cat = row_data.get('Category')
        if cat in category_colors:
            cell = ws_bestof.cell(row=excel_row, column=1)
            cell.fill = PatternFill(start_color=category_colors[cat], end_color=category_colors[cat], fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
    
    # === ARKUSZ 4: PROFILES ===
    ws_profiles = wb.create_sheet("PROFILES")
    style_sheet(ws_profiles, profiles_df, highlight_top=False)
    
    # Zapisz
    wb.save(output_path)
    logger.info(f"   ✅ Zapisano {len(consensus_df)} spółek")


def print_top_consensus(consensus_df: pd.DataFrame, n: int = 10):
    """Wyświetla TOP N w konsoli"""
    print(f"\n{'='*80}")
    print(f"CONSENSUS - TOP {n}")
    print(f"{'='*80}")
    print(f"{'Rank':<5} {'Ticker':<8} {'Signal':<8} {'Cover':<8} {'Elite':<7} {'Thesis'}")
    print("-" * 80)
    
    for _, row in consensus_df.head(n).iterrows():
        thesis = row['Investment_Thesis'][:40] + "..." if len(row['Investment_Thesis']) > 40 else row['Investment_Thesis']
        print(f"{row['Rank']:<5} {row['Ticker']:<8} {row['Signal_Strength']:<8} {row['Coverage']:<8} "
              f"{row['Elite_Score']:<7} {thesis}")


def print_ticker_details(companies: Dict[str, CompanySignals], ticker: str):
    """Wyświetla szczegóły dla spółki"""
    ticker = ticker.upper()
    if ticker not in companies:
        print(f"\n❌ Nie znaleziono spółki: {ticker}")
        return
    
    signals = companies[ticker]
    
    print(f"\n{'='*60}")
    print(f"PROFIL: {ticker}")
    print(f"{'='*60}")
    print(f"\n📊 Signal Strength: {signals.calculate_signal_strength():.1f}")
    print(f"📈 Coverage: {signals.coverage}/{len(MODEL_THEMES)} modeli")
    print(f"🏆 Elite Score: {signals.elite_score} (TOP5: {signals.top5_count}x, TOP10: {signals.top10_count}x)")
    print(f"🚩 Flagi: {''.join(f'[{f}]' for f in sorted(signals.unique_flags))}")
    print(f"⚠️  Warnings: {signals.warning_count}")
    
    print(f"\n💡 INVESTMENT THESIS:")
    print(f"   {signals.generate_thesis()}")
    
    print(f"\n📋 WYSTĄPIENIA W MODELACH:")
    for model, data in sorted(signals.appearances.items(), key=lambda x: x[1]['rank']):
        print(f"   • {model:<30} Rank: #{data['rank']:<3} Score: {data['score']:.1f} Flags: {data['flags']}")
    
    print(f"\n📊 SIŁA W KATEGORIACH:")
    for cat, strength in sorted(signals.get_category_strength().items(), key=lambda x: -x[1]):
        bar = "█" * int(strength) + "░" * (5 - int(strength))
        print(f"   {cat:<12} {bar} {strength:.1f}")


def show_status():
    """Pokazuje status systemu"""
    print(f"\n{'='*60}")
    print("GPW SCREENER - Status")
    print(f"{'='*60}")
    
    if os.path.exists(LATEST_FILE):
        mtime = os.path.getmtime(LATEST_FILE)
        mod_time = datetime.fromtimestamp(mtime)
        print(f"\n📊 Wyniki modeli: {mod_time.strftime('%Y-%m-%d %H:%M')}")
        print(f"   {LATEST_FILE}")
    else:
        print("\n⚠️  Brak wyników modeli. Uruchom: python run.py")
    
    if os.path.exists(OUTPUT_FILE):
        mtime = os.path.getmtime(OUTPUT_FILE)
        mod_time = datetime.fromtimestamp(mtime)
        print(f"\n📈 Wyniki ostateczne: {mod_time.strftime('%Y-%m-%d %H:%M')}")
        print(f"   {OUTPUT_FILE}")
    else:
        print("\n⚠️  Brak wyników ostatecznych. Uruchom: python app.py")


def main():
    parser = argparse.ArgumentParser(description='GPW Screener - Signal Aggregation')
    parser.add_argument('--status', action='store_true', help='Status systemu')
    parser.add_argument('--top', type=int, default=0, help='Pokaż TOP N w konsoli')
    parser.add_argument('--ticker', type=str, help='Szczegóły dla spółki')
    parser.add_argument('--no-save', action='store_true', help='Nie zapisuj do pliku')
    args = parser.parse_args()
    
    # Status
    if args.status:
        show_status()
        return
    
    # Wczytaj wyniki
    results = load_results()
    if not results:
        return
    
    # Agreguj sygnały
    logger.info("\n🔄 Agreguję sygnały...")
    companies = aggregate_signals(results)
    logger.info(f"   ✅ Znaleziono {len(companies)} unikalnych spółek")
    
    # Szczegóły dla tickera
    if args.ticker:
        print_ticker_details(companies, args.ticker)
        return
    
    # Twórz DataFrames
    logger.info("\n📊 Tworzę rankingi...")
    consensus_df = create_consensus_df(companies)
    heatmap_df = create_flag_heatmap(companies, top_n=30)
    best_of_df = create_best_of(companies)
    profiles_df = create_profiles(companies, results)
    
    # Zapisz Excel
    if not args.no_save:
        save_excel(consensus_df, heatmap_df, best_of_df, profiles_df, OUTPUT_FILE)
    
    # Pokaż TOP N
    top_n = args.top if args.top > 0 else 15
    print_top_consensus(consensus_df, top_n)
    
    # Podsumowanie
    print(f"\n{'='*80}")
    print("PODSUMOWANIE")
    print(f"{'='*80}")
    print(f"📊 Modeli: {len(results)}")
    print(f"📈 Spółek (unikalne): {len(companies)}")
    print(f"🏆 Spółki w TOP5 wielu modeli: {sum(1 for c in companies.values() if c.top5_count >= 2)}")
    print(f"🚩 Spółki z warning: {sum(1 for c in companies.values() if c.warning_count > 0)}")
    
    if not args.no_save:
        print(f"\n✅ Wyniki zapisane: {OUTPUT_FILE}")
    
    print(f"\n{'='*80}")
    print("✅ Gotowe!")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()
