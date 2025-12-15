#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
GPW SCREENER - Run All Models
================================================================================

Uruchamia wszystkie aktywne skanery i agreguje wyniki.

UŻYCIE:
    python run.py              # Uruchom wszystkie
    python run.py --list       # Lista skanerów
    python run.py --only qg    # Tylko Quality Growth
================================================================================
"""

import os
import sys
import shutil
import logging
import argparse
import importlib.util
from datetime import datetime
from typing import Dict, List, Optional, Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ścieżki
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SKANERY_DIR = os.path.join(BASE_DIR, "skanery")
DANE_DIR = os.path.join(BASE_DIR, "dane")
MAIN_DIR = os.path.join(BASE_DIR, "main")
ARCHIVE_DIR = os.path.join(MAIN_DIR, "archive")
LOGS_DIR = os.path.join(BASE_DIR, "logs")

sys.path.insert(0, SKANERY_DIR)


def setup_logging(log_to_file: bool = True) -> logging.Logger:
    """Konfiguruje logging"""
    logger = logging.getLogger("gpw_screener")
    logger.setLevel(logging.INFO)
    
    # Format
    formatter = logging.Formatter(
        '%(asctime)s [%(levelname)s] %(name)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Console handler
    console = logging.StreamHandler()
    console.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(console)
    
    # File handler
    if log_to_file:
        os.makedirs(LOGS_DIR, exist_ok=True)
        timestamp = datetime.now().strftime('%Y-%m-%d')
        log_file = os.path.join(LOGS_DIR, f"run_{timestamp}.log")
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
    
    return logger


def load_global_config() -> dict:
    """Wczytuje globalną konfigurację"""
    config_path = os.path.join(BASE_DIR, "config.yaml")
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f) or {}
    return {}


def discover_scanners() -> List[Dict[str, Any]]:
    """Automatycznie odkrywa wszystkie skanery"""
    scanners = []
    
    for item in sorted(os.listdir(SKANERY_DIR)):
        scanner_dir = os.path.join(SKANERY_DIR, item)
        model_file = os.path.join(scanner_dir, "model.py")
        config_file = os.path.join(scanner_dir, "config.yaml")
        
        if not os.path.isdir(scanner_dir) or not os.path.exists(model_file):
            continue
        
        # Wczytaj config skanera
        scanner_config = {}
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                scanner_config = yaml.safe_load(f) or {}
        
        # Sprawdź czy aktywny
        if not scanner_config.get('aktywny', True):
            continue
        
        # Ścieżka do danych
        data_filename = scanner_config.get('dane', 'dane.txt')
        data_path = os.path.join(DANE_DIR, data_filename)
        
        scanners.append({
            'id': item,
            'name': scanner_config.get('nazwa', item),
            'description': scanner_config.get('opis', ''),
            'model_path': model_file,
            'config_path': config_file,
            'data_path': data_path,
            'config': scanner_config
        })
    
    return scanners


def load_scanner_class(model_path: str):
    """Dynamicznie ładuje klasę skanera"""
    spec = importlib.util.spec_from_file_location("model", model_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    
    for name in dir(module):
        obj = getattr(module, name)
        if isinstance(obj, type) and name.endswith('Scanner') and name != 'BaseScanner':
            return obj
    
    return None


def run_scanners(scanner_filter: Optional[List[str]] = None) -> Dict[str, Any]:
    """Uruchamia skanery i zwraca wyniki"""
    from base import load_data
    
    logger = logging.getLogger("gpw_screener")
    results = {}
    scanners_info = discover_scanners()
    
    logger.info("=" * 60)
    logger.info("GPW SCREENER - Uruchamiam modele")
    logger.info("=" * 60)
    logger.info(f"Znaleziono {len(scanners_info)} aktywnych skanerów\n")
    
    for scanner_info in scanners_info:
        scanner_id = scanner_info['id']
        
        # Filtrowanie
        if scanner_filter and scanner_id not in scanner_filter:
            continue
        
        logger.info(f"📊 {scanner_info['name']}...")
        
        # Sprawdź dane
        if not os.path.exists(scanner_info['data_path']):
            logger.warning(f"   ⚠️  Brak pliku danych: {scanner_info['data_path']}")
            continue
        
        try:
            # Załaduj dane
            df = load_data(scanner_info['data_path'])
            if df is None:
                continue
            
            # Załaduj i uruchom skaner
            ScannerClass = load_scanner_class(scanner_info['model_path'])
            if ScannerClass is None:
                logger.error(f"   ❌ Nie znaleziono klasy skanera")
                continue
            
            scanner = ScannerClass(scanner_info['config_path'])
            scanner.run(df)
            
            results[scanner_id] = {
                'scanner': scanner,
                'data': scanner.results,
                'columns': scanner.get_output_columns(),
                'info': scanner_info
            }
            
            logger.info(f"   ✅ Przetworzono {len(scanner.results)} spółek")
            
        except Exception as e:
            logger.error(f"   ❌ Błąd: {e}")
            import traceback
            logger.debug(traceback.format_exc())
    
    return results


def create_excel(results: Dict[str, Any], output_path: str):
    """Tworzy plik Excel z wynikami"""
    logger = logging.getLogger("gpw_screener")
    
    wb = Workbook()
    
    # Style
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    top5_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    top10_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # === ARKUSZ PODSUMOWANIA ===
    ws_summary = wb.active
    ws_summary.title = "PODSUMOWANIE"
    
    ws_summary['A1'] = "GPW SCREENER - Wyniki"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary['A2'].font = Font(italic=True, color="666666")
    
    # Tabela podsumowania
    summary_headers = ['Model', 'Spółek', 'TOP 1', 'TOP 2', 'TOP 3', 'TOP 4', 'TOP 5']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    row = 5
    for scanner_id, result in results.items():
        scanner = result['scanner']
        df = result['data']
        
        ws_summary.cell(row=row, column=1, value=scanner.name).border = thin_border
        ws_summary.cell(row=row, column=2, value=len(df)).border = thin_border
        
        top5 = df.head(5)
        for i, (_, r) in enumerate(top5.iterrows()):
            cell = ws_summary.cell(row=row, column=3+i, value=f"{r['Ticker']} ({r['Total']:.0f})")
            cell.border = thin_border
            if i == 0:
                cell.font = Font(bold=True)
        
        row += 1
    
    # Szerokości kolumn
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 10
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws_summary.column_dimensions[col].width = 15
    
    # === ARKUSZE MODELI ===
    for scanner_id, result in results.items():
        scanner = result['scanner']
        df = result['data']
        columns = result['columns']
        
        # Filtruj kolumny
        available_cols = [c for c in columns if c in df.columns]
        
        # Nazwa arkusza (max 31 znaków)
        sheet_name = scanner.name[:31].replace('/', '-')
        ws = wb.create_sheet(sheet_name)
        
        # Nagłówki
        for col_idx, col_name in enumerate(available_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Dane
        for row_idx, row_data in df.iterrows():
            excel_row = row_idx + 2
            
            for col_idx, col_name in enumerate(available_cols, 1):
                value = row_data.get(col_name, '')
                
                # Formatowanie
                if isinstance(value, float):
                    if col_name in ['Debt_Ratio', 'Asset_Coverage', 'P_E', 'P_BV']:
                        value = round(value, 2)
                    else:
                        value = round(value, 1)
                
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = thin_border
                
                # Kolorowanie
                if row_idx < 5:
                    cell.fill = top5_fill
                elif row_idx < 10:
                    cell.fill = top10_fill
        
        # Szerokości
        for col_idx, col_name in enumerate(available_cols, 1):
            width = max(10, len(str(col_name)) + 2)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    wb.save(output_path)
    logger.info(f"✅ Zapisano: {output_path}")


def archive_old_results():
    """Archiwizuje poprzednie wyniki"""
    logger = logging.getLogger("gpw_screener")
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    
    for filename in os.listdir(MAIN_DIR):
        if filename.startswith("wyniki_") and filename.endswith(".xlsx"):
            if filename == "wyniki_latest.xlsx":
                continue
            src = os.path.join(MAIN_DIR, filename)
            dst = os.path.join(ARCHIVE_DIR, filename)
            if not os.path.exists(dst):
                shutil.move(src, dst)
                logger.debug(f"Zarchiwizowano: {filename}")


def print_summary(results: Dict[str, Any]):
    """Wyświetla podsumowanie"""
    logger = logging.getLogger("gpw_screener")
    
    logger.info("\n" + "=" * 70)
    logger.info("PODSUMOWANIE - TOP 5")
    logger.info("=" * 70)
    
    for scanner_id, result in results.items():
        scanner = result['scanner']
        df = result['data']
        
        logger.info(f"\n📈 {scanner.name}")
        logger.info("-" * 50)
        
        for _, row in df.head(5).iterrows():
            flags = row.get('Flags', '')
            logger.info(f"   {row['Rank']:>2}. {row['Ticker']:<8} {flags:<14} Score: {row['Total']:.1f}")


def main():
    parser = argparse.ArgumentParser(description='GPW Screener - Run All Models')
    parser.add_argument('--list', action='store_true', help='Lista dostępnych skanerów')
    parser.add_argument('--only', nargs='+', help='Uruchom tylko wybrane skanery')
    parser.add_argument('--no-archive', action='store_true', help='Nie archiwizuj starych wyników')
    parser.add_argument('--debug', action='store_true', help='Debug mode')
    args = parser.parse_args()
    
    # Setup
    logger = setup_logging()
    if args.debug:
        logger.setLevel(logging.DEBUG)
    
    # Lista skanerów
    if args.list:
        scanners = discover_scanners()
        print("\nDostępne skanery:")
        print("-" * 50)
        for s in scanners:
            print(f"  {s['id']:<20} {s['name']}")
        print()
        return
    
    # Utwórz foldery
    os.makedirs(MAIN_DIR, exist_ok=True)
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    
    # Archiwizuj stare
    if not args.no_archive:
        archive_old_results()
    
    # Uruchom skanery
    results = run_scanners(args.only)
    
    if not results:
        logger.error("\n❌ Brak wyników do zapisania")
        return
    
    # Zapisz Excel z timestampem
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
    output_file = os.path.join(MAIN_DIR, f"wyniki_{timestamp}.xlsx")
    create_excel(results, output_file)
    
    # Utwórz/nadpisz "latest"
    latest_file = os.path.join(MAIN_DIR, "wyniki_latest.xlsx")
    shutil.copy(output_file, latest_file)
    logger.info(f"✅ Latest: {latest_file}")
    
    # Podsumowanie
    print_summary(results)
    
    logger.info(f"\n{'='*70}")
    logger.info("✅ Gotowe!")
    logger.info(f"{'='*70}\n")


if __name__ == "__main__":
    main()
